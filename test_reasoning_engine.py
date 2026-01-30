#!/usr/bin/env python3
"""
Simplified Test Version of Reasoning Engine
Tests core logic WITHOUT Google API dependencies
"""

import argparse
import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('test_app.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

SHORTLIST_THRESHOLD = 69
REJECT_THRESHOLD = 50

def parse_json(file_path):
    """Parse candidate data from JSON file."""
    logger.info(f"Parsing candidate data from: {file_path}")
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"JSON file not found: {file_path}")
    
    with open(file_path, 'r', encoding='utf-8') as f:
        candidates = json.load(f)
    
    logger.info(f"Successfully parsed {len(candidates)} candidates")
    return candidates

def determine_decision(similarity_score):
    """Determine hiring decision based on similarity score."""
    if similarity_score >= SHORTLIST_THRESHOLD:
        return "Shortlisted"
    elif similarity_score >= REJECT_THRESHOLD:
        return "Maybe"
    else:
        return "Rejected"

def generate_reasoning(similarity_score, decision):
    """Generate professional reasoning text."""
    score_formatted = f"{similarity_score:.1f}%"
    
    reasoning_templates = {
        "Shortlisted": (
            f"Strong match with job requirements. Similarity score: {score_formatted}. "
            f"Candidate demonstrates excellent alignment with required skills and qualifications. "
            f"The high similarity score indicates strong compatibility with the position requirements, "
            f"making this candidate a prime choice for further interview rounds."
        ),
        "Maybe": (
            f"Moderate match with requirements. Similarity score: {score_formatted}. "
            f"Candidate shows partial alignment with job requirements. "
            f"Recommended for further review and interview assessment to evaluate potential fit. "
            f"Additional screening may reveal transferable skills or growth potential."
        ),
        "Rejected": (
            f"Insufficient match with job requirements. Similarity score: {score_formatted}. "
            f"Candidate's profile does not meet the minimum threshold for this position. "
            f"The similarity score indicates limited alignment with required competencies and qualifications. "
            f"Consider for alternative positions or future opportunities with different requirements."
        )
    }
    
    return reasoning_templates.get(decision, f"Similarity score: {score_formatted}")

def process_candidates(candidates):
    """Process all candidates without Google API calls."""
    logger.info(f"Processing {len(candidates)} candidates...")
    processed_candidates = []
    stats = {"Shortlisted": 0, "Maybe": 0, "Rejected": 0}
    
    for idx, candidate in enumerate(candidates, 1):
        logger.info(f"Processing candidate {idx}/{len(candidates)}: {candidate['candidate_name']}")
        
        similarity_score = float(candidate['similarity_score'])
        decision = determine_decision(similarity_score)
        stats[decision] += 1
        reasoning = generate_reasoning(similarity_score, decision)
        
        # Mock resume URL (since we're not uploading)
        resume_file = candidate['resume_file_path']
        resume_url = f"Local file: {resume_file}"
        
        processed_candidate = {
            'candidate_name': candidate['candidate_name'],
            'email': candidate['email'],
            'phone': candidate['phone'],
            'experience_years': candidate['experience_years'],
            'skills': ', '.join(candidate['skills']) if isinstance(candidate['skills'], list) else candidate['skills'],
            'education': candidate['education'],
            'similarity_score': similarity_score,
            'decision': decision,
            'reasoning': reasoning,
            'resume_url': resume_url,
            'processed_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        processed_candidates.append(processed_candidate)
        logger.info(f"✓ {candidate['candidate_name']}: {decision} ({similarity_score:.1f}%)")
    
    logger.info(f"Processing complete. Stats: {stats}")
    return processed_candidates, stats

def create_excel_report(candidates, output_path):
    """Create formatted Excel report."""
    logger.info(f"Creating Excel report: {output_path}")
    
    df = pd.DataFrame(candidates)
    
    # Reorder columns
    column_order = [
        'candidate_name', 'email', 'phone', 'experience_years',
        'similarity_score', 'decision', 'reasoning', 'resume_url',
        'skills', 'education', 'processed_date'
    ]
    df = df[column_order]
    
    # Rename columns
    df.columns = [
        'Candidate Name', 'Email', 'Phone', 'Experience (Years)',
        'Similarity Score (%)', 'Decision', 'Reasoning', 'Resume URL',
        'Skills', 'Education', 'Processed Date'
    ]
    
    # Save to Excel
    df.to_excel(output_path, index=False, engine='openpyxl')
    
    # Apply formatting
    format_excel(output_path)
    
    logger.info(f"Excel report created successfully: {output_path}")
    return output_path

def format_excel(file_path):
    """Apply professional formatting to Excel report."""
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define colors
    decision_colors = {
        'Shortlisted': 'C6EFCE',
        'Maybe': 'FFEB9C',
        'Rejected': 'FFC7CE'
    }
    
    # Header formatting
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'
    
    # Format data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        decision_cell = row[5]
        score_cell = row[4]
        
        decision_cell.alignment = Alignment(horizontal='center', vertical='center')
        score_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        decision_value = decision_cell.value
        if decision_value in decision_colors:
            decision_cell.fill = PatternFill(
                start_color=decision_colors[decision_value],
                end_color=decision_colors[decision_value],
                fill_type='solid'
            )
        
        reasoning_cell = row[6]
        reasoning_cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Column widths
    column_widths = {
        'A': 20, 'B': 25, 'C': 15, 'D': 12, 'E': 15, 'F': 12,
        'G': 60, 'H': 40, 'I': 35, 'J': 25, 'K': 20
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 60
    
    ws.row_dimensions[1].height = 30
    
    wb.save(file_path)
    logger.info("Excel formatting applied successfully")

def main():
    parser = argparse.ArgumentParser(description='Test Reasoning Engine (No Google APIs)')
    parser.add_argument('--input', '-i', required=True, help='Input JSON file')
    parser.add_argument('--output', '-o', required=True, help='Output Excel file')
    args = parser.parse_args()
    
    logger.info("=" * 80)
    logger.info("TEST REASONING ENGINE - STARTED")
    logger.info("=" * 80)
    
    try:
        # Parse JSON
        logger.info("\n[STEP 1/3] Parsing candidate data...")
        candidates = parse_json(args.input)
        
        # Process candidates
        logger.info("\n[STEP 2/3] Processing candidates...")
        processed_candidates, stats = process_candidates(candidates)
        
        # Create Excel report
        logger.info("\n[STEP 3/3] Generating Excel report...")
        excel_path = create_excel_report(processed_candidates, args.output)
        
        # Summary
        logger.info("\n" + "=" * 80)
        logger.info("SCREENING SUMMARY")
        logger.info("=" * 80)
        logger.info(f"Total Candidates Processed: {len(processed_candidates)}")
        logger.info(f"✓ Shortlisted: {stats['Shortlisted']}")
        logger.info(f"⚠ Maybe: {stats['Maybe']}")
        logger.info(f"✗ Rejected: {stats['Rejected']}")
        logger.info(f"\nExcel Report: {excel_path}")
        logger.info("=" * 80)
        
    except Exception as e:
        logger.error(f"Error: {e}", exc_info=True)
        sys.exit(1)

if __name__ == '__main__':
    main()
