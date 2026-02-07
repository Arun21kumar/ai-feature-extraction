#!/usr/bin/env python3
"""
Automated Candidate Screening Reasoning Engine

This module provides intelligent candidate screening based on similarity scores,
automated resume upload to Google Drive, Excel report generation, and email
notification via Gmail API.

Author: AI Feature Extraction Team
Date: January 2026
"""

import argparse
import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from google_drive_service import GoogleDriveService
from gmail_service import GmailService

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()


class CandidateScreeningEngine:
    """
    Main reasoning engine for automated candidate screening.
    
    This class orchestrates the complete workflow:
    1. Parse candidate JSON data
    2. Apply decision logic based on similarity scores
    3. Upload resumes to Google Drive
    4. Generate Excel reports with reasoning
    5. Send email notifications with reports
    """
    
    def __init__(self):
        """Initialize the screening engine with environment configurations."""
        self.shortlist_threshold = float(os.getenv('SHORTLIST_THRESHOLD', 69))
        self.reject_threshold = float(os.getenv('REJECT_THRESHOLD', 50))
        
        # Parse comma-separated list of HR email recipients
        hr_email_str = os.getenv('HR_EMAIL_RECIPIENT', '')
        self.hr_recipients = [email.strip() for email in hr_email_str.split(',') if email.strip()]
        if not self.hr_recipients:
            raise ValueError("HR_EMAIL_RECIPIENT environment variable is not set or is empty.")

        self.drive_folder_id = os.getenv('DRIVE_FOLDER_ID')
        
        # Add a cache to store links of already uploaded files
        self.upload_cache = {}
        
        # Get path to OAuth client secrets
        self.credentials_path = os.getenv('GOOGLE_CREDENTIALS_PATH')
        if not self.credentials_path:
            raise ValueError("GOOGLE_CREDENTIALS_PATH environment variable not set.")

        # Initialize Google services
        try:
            # Both services will now use the same OAuth credentials and token file
            self.drive_service = GoogleDriveService(credentials_path=self.credentials_path)
            self.gmail_service = GmailService(credentials_path=self.credentials_path)
            logger.info("Google services initialized successfully")
        except Exception as e:
            logger.error(f"Failed to initialize Google services: {e}")
            logger.error(
                "Please ensure your 'credentials.json' is correctly placed and "
                "that you have authorized the application on first run."
            )
            raise
    
    def parse_json(self, file_path: str) -> List[Dict]:
        """
        Parse candidate data from JSON file.
        
        Args:
            file_path: Path to the JSON file containing candidate data
            
        Returns:
            List of candidate dictionaries
            
        Raises:
            FileNotFoundError: If JSON file doesn't exist
            json.JSONDecodeError: If JSON format is invalid
            ValueError: If required fields are missing
        """
        logger.info(f"Parsing candidate data from: {file_path}")
        
        if not os.path.exists(file_path):
            logger.error(f"JSON file not found: {file_path}")
            raise FileNotFoundError(f"JSON file not found: {file_path}")
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                candidates = json.load(f)
            
            # Validate JSON structure
            if not isinstance(candidates, list):
                raise ValueError("JSON must contain a list of candidates")
            
            required_fields = [
                'candidate_name', 'email', 'phone', 'experience_years',
                'skills', 'education', 'similarity_score', 'resume_file_path'
            ]
            
            for idx, candidate in enumerate(candidates):
                missing_fields = [field for field in required_fields if field not in candidate]
                if missing_fields:
                    logger.error(f"Candidate {idx} missing fields: {missing_fields}")
                    raise ValueError(f"Candidate {idx} missing required fields: {missing_fields}")
            
            logger.info(f"Successfully parsed {len(candidates)} candidates")
            return candidates
            
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON format: {e}")
            raise
        except Exception as e:
            logger.error(f"Error parsing JSON: {e}")
            raise
    
    def determine_decision(self, similarity_score: float) -> str:
        """
        Determine hiring decision based on similarity score.
        
        Args:
            similarity_score: Candidate's similarity score (0-100)
            
        Returns:
            Decision string: "Shortlisted", "Maybe", or "Rejected"
        """
        if similarity_score >= self.shortlist_threshold:
            return "Shortlisted"
        elif similarity_score >= self.reject_threshold:
            return "Maybe"
        else:
            return "Rejected"
    
    def generate_reasoning(self, similarity_score: float, decision: str, section_scores: dict = None) -> str:
        """
        Generate professional reasoning text based on decision outcome.
        
        Args:
            similarity_score: Candidate's similarity score
            decision: Hiring decision (Shortlisted/Maybe/Rejected)
            section_scores: Optional dict with section-wise scores (summary, skills, responsibilities, certifications)
            
        Returns:
            Detailed reasoning text explaining the decision
        """
        score_formatted = f"{similarity_score:.1f}%"
        
        # Build detailed section analysis if available
        section_analysis = ""
        if section_scores and isinstance(section_scores, dict):
            highlights = []
            concerns = []
            
            for section, score in section_scores.items():
                if section == 'overall_score':
                    continue
                if score >= 70:
                    highlights.append(f"{section.capitalize()}: {score:.1f}%")
                elif score < 50:
                    concerns.append(f"{section.capitalize()}: {score:.1f}%")
            
            if highlights:
                section_analysis = f" Strong areas: {', '.join(highlights)}."
            if concerns:
                section_analysis += f" Areas for review: {', '.join(concerns)}."
        
        reasoning_templates = {
            "Shortlisted": (
                f"Strong match with job requirements. Overall similarity score: {score_formatted}.{section_analysis} "
                f"Candidate demonstrates excellent alignment with required skills and qualifications. "
                f"The high similarity score indicates strong compatibility with the position requirements, "
                f"making this candidate a prime choice for further interview rounds."
            ),
            "Maybe": (
                f"Moderate match with requirements. Overall similarity score: {score_formatted}.{section_analysis} "
                f"Candidate shows partial alignment with job requirements. "
                f"Recommended for further review and interview assessment to evaluate potential fit. "
                f"Additional screening may reveal transferable skills or growth potential."
            ),
            "Rejected": (
                f"Insufficient match with job requirements. Overall similarity score: {score_formatted}.{section_analysis} "
                f"Candidate's profile does not meet the minimum threshold for this position. "
                f"The similarity score indicates limited alignment with required competencies and qualifications. "
                f"Consider for alternative positions or future opportunities with different requirements."
            )
        }
        
        return reasoning_templates.get(decision, f"Overall similarity score: {score_formatted}")
    
    def upload_resume_to_drive(self, file_path: str, candidate_name: str) -> str:
        """
        Upload resume file to Google Drive and return shareable link.
        
        Args:
            file_path: Local path to resume file (PDF/DOCX)
            candidate_name: Name of the candidate (for file naming)
            
        Returns:
            Google Drive shareable link
            
        Raises:
            FileNotFoundError: If resume file doesn't exist
            Exception: If upload fails
        """
        logger.info(f"Uploading resume for {candidate_name}: {file_path}")
        
        if not os.path.exists(file_path):
            logger.error(f"Resume file not found: {file_path}")
            raise FileNotFoundError(f"Resume file not found: {file_path}")
        
        try:
            # Upload to Google Drive
            drive_link, drive_id = self.drive_service.upload_file(
                file_path=file_path,
                folder_id=self.drive_folder_id,
                file_name=f"{candidate_name}_Resume{Path(file_path).suffix}"
            )
            
            logger.info(f"Resume uploaded successfully for {candidate_name}")
            # Return both link and ID
            return drive_link, drive_id
            
        except Exception as e:
            logger.error(f"Failed to upload resume for {candidate_name}: {e}")
            # Return placeholder link on failure to not block the entire process
            return f"Upload failed: {str(e)[:50]}", None
    
    def process_candidates(self, candidates: List[Dict]) -> Tuple[List[Dict], Dict[str, int]]:
        """
        Process all candidates: determine decisions, generate reasoning, upload resumes.
        
        Args:
            candidates: List of candidate dictionaries
            
        Returns:
            Tuple of (processed_candidates, summary_statistics)
        """
        logger.info(f"Processing {len(candidates)} candidates...")
        processed_candidates = []
        stats = {"Shortlisted": 0, "Maybe": 0, "Rejected": 0}
        
        for idx, candidate in enumerate(candidates, 1):
            try:
                logger.info(f"Processing candidate {idx}/{len(candidates)}: {candidate['candidate_name']}")
                
                # Extract similarity score
                similarity_score = float(candidate['similarity_score'])
                
                # Determine decision
                decision = self.determine_decision(similarity_score)
                stats[decision] += 1
                
                # Get section scores if available
                section_scores = candidate.get('section_scores', {})
                
                # Generate reasoning with section scores
                reasoning = self.generate_reasoning(similarity_score, decision, section_scores)
                
                # --- CACHING & UPLOAD LOGIC ---
                resume_path = candidate.get('resume_file_path')
                resume_url = "Resume Not Provided"  # Default value

                if resume_path:  # Only proceed if a path is provided
                    # Check if the resume has already been uploaded in this run
                    if resume_path in self.upload_cache:
                        resume_url = self.upload_cache[resume_path]['link']
                        logger.info(f"Using cached resume link for {resume_path}")
                    else:
                        # If not cached, upload the resume
                        try:
                            uploaded_url, uploaded_id = self.upload_resume_to_drive(
                                resume_path,
                                candidate['candidate_name']
                            )
                            resume_url = uploaded_url
                            # Save the new link and ID to the cache
                            if uploaded_id:
                                self.upload_cache[resume_path] = {'link': uploaded_url, 'id': uploaded_id}
                        except Exception as e:
                            logger.error(f"Resume upload failed for {candidate['candidate_name']}: {e}")
                            resume_url = "Upload failed - see logs"
                else:
                    logger.warning(f"No resume_file_path provided for {candidate['candidate_name']}. Skipping upload.")
                # --- END CACHING & UPLOAD LOGIC ---

                # Compile processed candidate data
                processed_candidate = {
                    'candidate_name': candidate['candidate_name'],
                    'email': candidate['email'],
                    'phone': candidate['phone'],
                    'experience_years': candidate['experience_years'],
                    'skills': ', '.join(candidate['skills']) if isinstance(candidate['skills'], list) else candidate['skills'],
                    'education': candidate['education'],
                    'similarity_score': similarity_score,
                    'decision': decision,
                    'reasoning': reasoning,
                    'resume_url': resume_url,
                    'processed_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                processed_candidates.append(processed_candidate)
                logger.info(f"✓ {candidate['candidate_name']}: {decision} ({similarity_score:.1f}%)")
                
            except Exception as e:
                logger.error(f"Error processing candidate {candidate.get('candidate_name', 'Unknown')}: {e}")
                # Continue processing other candidates
                continue
        
        logger.info(f"Processing complete. Stats: {stats}")
        return processed_candidates, stats
    
    def create_excel_report(self, candidates: List[Dict], output_path: str) -> str:
        """
        Create formatted Excel report with candidate screening results.
        
        Args:
            candidates: List of processed candidate dictionaries
            output_path: Path where Excel file should be saved
            
        Returns:
            Path to created Excel file
        """
        logger.info(f"Creating Excel report: {output_path}")
        
        try:
            # Create DataFrame
            df = pd.DataFrame(candidates)
            
            # Reorder columns for better readability
            column_order = [
                'candidate_name', 'email', 'phone', 'experience_years',
                'similarity_score', 'decision', 'reasoning', 'resume_url',
                'skills', 'education', 'processed_date'
            ]
            df = df[column_order]
            
            # Rename columns for display
            df.columns = [
                'Candidate Name', 'Email', 'Phone', 'Experience (Years)',
                'Similarity Score (%)', 'Decision', 'Reasoning', 'Resume URL',
                'Skills', 'Education', 'Processed Date'
            ]
            
            # Save to Excel
            df.to_excel(output_path, index=False, engine='openpyxl')
            
            # Apply formatting
            self._format_excel(output_path)
            
            logger.info(f"Excel report created successfully: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Failed to create Excel report: {e}")
            raise
    
    def _format_excel(self, file_path: str):
        """
        Apply professional formatting to Excel report.
        
        Args:
            file_path: Path to Excel file to format
        """
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Define colors for decisions
            decision_colors = {
                'Shortlisted': 'C6EFCE',  # Light green
                'Maybe': 'FFEB9C',        # Light yellow
                'Rejected': 'FFC7CE'      # Light red
            }
            
            # Header formatting
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Freeze top row
            ws.freeze_panes = 'A2'
            
            # Format data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                # Center align Decision and Similarity Score columns
                decision_cell = row[5]  # Decision column (F)
                score_cell = row[4]     # Similarity Score column (E)
                
                decision_cell.alignment = Alignment(horizontal='center', vertical='center')
                score_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Apply color based on decision
                decision_value = decision_cell.value
                if decision_value in decision_colors:
                    decision_cell.fill = PatternFill(
                        start_color=decision_colors[decision_value],
                        end_color=decision_colors[decision_value],
                        fill_type='solid'
                    )
                
                # Wrap text in reasoning column
                reasoning_cell = row[6]  # Reasoning column (G)
                reasoning_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Auto-adjust column widths
            column_widths = {
                'A': 20,  # Candidate Name
                'B': 25,  # Email
                'C': 15,  # Phone
                'D': 12,  # Experience
                'E': 15,  # Similarity Score
                'F': 12,  # Decision
                'G': 60,  # Reasoning
                'H': 40,  # Resume URL
                'I': 35,  # Skills
                'J': 25,  # Education
                'K': 20,  # Processed Date
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Set row heights
            for row in ws.iter_rows(min_row=2):
                ws.row_dimensions[row[0].row].height = 60
            
            ws.row_dimensions[1].height = 30
            
            wb.save(file_path)
            logger.info("Excel formatting applied successfully")
            
        except Exception as e:
            logger.error(f"Failed to format Excel file: {e}")
            # Don't raise - formatting is optional
    
    def send_email_with_report(
        self,
        excel_path: str,
        recipient_email: str,
        summary_stats: Dict[str, int],
        total_candidates: int
    ):
        """
        Send email with Excel report attachment and summary.
        
        Args:
            excel_path: Path to Excel report file
            recipient_email: Email address of recipient
            summary_stats: Dictionary with counts of Shortlisted/Maybe/Rejected
            total_candidates: Total number of candidates processed
        """
        logger.info(f"Sending email report to {recipient_email}")
        
        try:
            # Import email template
            from email_template import generate_email_html
            
            # Generate email content
            subject = f"Candidate Screening Report - {datetime.now().strftime('%B %d, %Y')}"
            
            html_body = generate_email_html(
                total_candidates=total_candidates,
                shortlisted=summary_stats.get('Shortlisted', 0),
                maybe=summary_stats.get('Maybe', 0),
                rejected=summary_stats.get('Rejected', 0)
            )
            
            # Send email via Gmail API
            self.gmail_service.send_email_with_attachment(
                to=recipient_email,
                subject=subject,
                html_body=html_body,
                attachment_path=excel_path
            )
            
            logger.info(f"Email sent successfully to {recipient_email}")
            
        except Exception as e:
            logger.error(f"Failed to send email: {e}")
            raise
    
    def run(self, input_json: str, output_excel: str):
        """
        Main orchestration method for the entire screening workflow.
        
        Args:
            input_json: Path to input JSON file with candidate data
            output_excel: Path where Excel report should be saved
        """
        logger.info("=" * 80)
        logger.info("AUTOMATED CANDIDATE SCREENING ENGINE - STARTED")
        logger.info("=" * 80)
        
        try:
            # Step 1: Parse JSON
            logger.info("\n[STEP 1/5] Parsing candidate data...")
            candidates = self.parse_json(input_json)
            
            # Step 2: Process candidates
            logger.info("\n[STEP 2/5] Processing candidates and uploading resumes...")
            processed_candidates, stats = self.process_candidates(candidates)
            
            if not processed_candidates:
                logger.error("No candidates were successfully processed")
                return
            
            # Step 3: Create Excel report
            logger.info("\n[STEP 3/5] Generating Excel report...")
            excel_path = self.create_excel_report(processed_candidates, output_excel)
            
            # Step 4: Send email
            logger.info("\n[STEP 4/5] Sending email notification...")
            if self.hr_recipients:
                # Convert list of recipients to a comma-separated string for the 'To' field
                recipient_str = ", ".join(self.hr_recipients)
                self.send_email_with_report(
                    excel_path=excel_path,
                    recipient_email=recipient_str,
                    summary_stats=stats,
                    total_candidates=len(candidates)
                )
            else:
                logger.warning("No HR email recipients configured. Skipping email notification.")
            
            # Step 5: Finalize
            logger.info("\n[STEP 5/5] Process complete.")
            logger.info("=" * 80)
            logger.info("AUTOMATED CANDIDATE SCREENING ENGINE - FINISHED")
            logger.info("=" * 80)
            
            if self.hr_recipients:
                logger.info(f"✅ Success! Report sent to {', '.join(self.hr_recipients)}")

        except Exception as e:
            logger.error(f"Screening engine failed: {e}")
            raise


def main():
    """Main entry point for command-line execution."""
    parser = argparse.ArgumentParser(
        description='Automated Candidate Screening Reasoning Engine',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python reasoning_engine.py --input candidates.json --output report.xlsx
  python reasoning_engine.py -i data/candidates.json -o output/screening_report.xlsx
        """
    )
    
    parser.add_argument(
        '--input', '-i',
        required=True,
        help='Path to input JSON file containing candidate data'
    )
    
    parser.add_argument(
        '--output', '-o',
        required=True,
        help='Path where Excel report should be saved'
    )
    
    args = parser.parse_args()
    
    # Validate inputs
    if not args.input.endswith('.json'):
        logger.error("Input file must be a JSON file")
        sys.exit(1)
    
    if not args.output.endswith('.xlsx'):
        logger.error("Output file must be an Excel file (.xlsx)")
        sys.exit(1)
    
    # Create output directory if it doesn't exist
    output_dir = os.path.dirname(args.output)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        logger.info(f"Created output directory: {output_dir}")
    
    # Run the screening engine
    try:
        engine = CandidateScreeningEngine()
        engine.run(input_json=args.input, output_excel=args.output)
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
